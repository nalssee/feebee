import locale
import csv
from openpyxl import load_workbook
from datetime import timedelta
import numpy as np
from itertools import zip_longest, accumulate, groupby
import pandas as pd
import ciso8601
import statsmodels.api as sm


def lag(cols, datecol, ns, add1fn=None, max_missings=10_000):
    """ create columns for lags and leads, ex) col_1, col_2n
    """
    cols = listify(cols)
    ns = listify(ns)

    def fn(rs):
        suffix = "_"
        rs.sort(key=lambda r: r[datecol])
        if add1fn:
            rs1 = [rs[0]]
            for r1, r2 in zip(rs, rs[1:]):
                d1, d2 = r1[datecol], r2[datecol]
                if d1 == d2:
                    raise ValueError("Duplicates for lead lags", r1, r2)
                if add1fn(d1) == d2:
                    rs1.append(r2)
                else:
                    cnt = 0
                    d1 = add1fn(d1)
                    while d1 != d2:
                        if cnt >= max_missings:
                            raise ValueError("Too many missings", r1, r2)
                        r0 = _empty(r1)
                        r0[datecol] = d1
                        rs1.append(r0)
                        d1 = add1fn(d1)
                        cnt += 1
                    rs1.append(r2)
        else:
            rs1 = rs

        r0 = rs1[0]
        rss = [rs1]
        for n in ns:
            if n >= 0:
                rss.append([_empty(r0)] * n + rs1)
            else:
                rss.append(rs1[(-n):] + [_empty(r0)] * (-n))

        result = []
        for xs in zip(*rss):
            x0 = xs[0]
            for n, x in zip(ns, xs[1:]):
                for c in cols:
                    strn = str(n) if n >=0 else str(-n) + 'n'
                    x0[c + suffix + strn] = x[c]
            result.append(x0)
        return result
    return fn


def where(pred, fn=None):
    """ Filter with pred before you apply fn to a list of rows.
        if fn is not given, simply filter with pred
    """
    if fn:
        def func(rs):
            return fn([r for r in rs if pred(r)])
        return func
    else:
        return lambda r: (r if pred(r) else None)


def affix(**kwargs):
    def fn(r):
        for k, v in kwargs.items():
            try:
                r[k] = v(r)
            except:
                r[k] = ''
        return r
    return fn


def read_date(date):
    return ciso8601.parse_datetime(date)


def add_date(date, n):
    """n means days for "%Y-%m-%d", months for "%Y-%m". Supports only 2 fmt.
    """
    # "1903-09"
    if len(date) == 7:
        y, m = date.split("-")
        n1 = int(y) * 12 + int(m) + n
        y1, m1 = n1 // 12, n1 % 12
        if m1 == 0:
            y1 -= 1
            m1 = 12
        return str(y1) + '-' + str(m1).zfill(2)
    # "1903-09-29"
    elif len(date) == 10:
        d = ciso8601.parse_datetime(date)
        d1 = d + timedelta(days=n)
        return d1.strftime("%Y-%m-%d")
    else:
        raise ValueError("Unsupported date format", date)


def ols(rs, y, xs, add_constant=True):
    xs = listify(xs)
    df = pd.DataFrame(rs)
    if add_constant:
        return sm.OLS(df[[y]], sm.add_constant(df[list(xs)])).fit()
    else:
        return sm.OLS(df[[y]], df[list(xs)]).fit()


def logit(rs, y, xs, add_constant=True):
    xs = listify(xs)
    df = pd.DataFrame(rs)
    if add_constant:
        # disp=0 suppress some infos regarding optimization
        return sm.Logit(df[[y]], sm.add_constant(df[list(xs)])).fit(disp=0)
    else:
        return sm.Logit(df[[y]], df[list(xs)]).fit(disp=0)


def chunk(rs, n, column=None):
    """
    Usage:
        |  chunk(rs, 3) => returns 3 rows about the same size
        |  chunk(rs, [0.3, 0.4, 0.3]) => returns 3 rows of 30%, 40%, 30%
        |  chunk(rs, [100, 500, 1000], 'col')
        |      => returns 4 rows with break points 100, 500, 1000 of 'col'
    """
    size = len(rs)
    if isinstance(n, int):
        start = 0
        result = []
        for i in range(1, n + 1):
            end = int((size * i) / n)
            # must yield anyway
            result.append(rs[start:end])
            start = end
        return result
    # n is a list of percentiles
    elif not column:
        # then it is a list of percentiles for each chunk
        assert sum(n) <= 1, f"Sum of percentils for chunks must be <= 1.0"
        ns = [int(x * size) for x in accumulate(n)]
        result = []
        for a, b in zip([0] + ns, ns):
            result.append(rs[a:b])
        return result
    # n is a list of break points
    else:
        rs.sort(key=lambda r: r[column])
        start, end = 0, 0
        result = []
        for bp in n:
            while (rs[end][column] < bp) and end < size:
                end += 1
            result.append(rs[start:end])
            start = end
        result.append(rs[end:])
        return result


def winsorize(rs, col, limit=0.01):
    """Winsorsize rows that are out of limits
    Args:
        |  col(str): column name.
        |  limit(float): for both sides respectably.
    returns rs
    """
    rs = [r for r in rs if isnum(r[col])]
    xs = [r[col] for r in rs]
    lower = np.percentile(xs, limit * 100)
    higher = np.percentile(xs, (1 - limit) * 100)
    for r in rs:
        if r[col] > higher:
            r[col] = higher
        elif r[col] < lower:
            r[col] = lower
    return rs


def truncate(rs, col, limit=0.01):
    """Truncate rows that are out of limits
    Args:
        |  col(str): column name
        |  limit(float): for both sides respectably.
    Returns self
    """
    rs = [r for r in rs if isnum(r[col])]
    xs = [r[col] for r in rs]
    lower = np.percentile(xs, limit * 100)
    higher = np.percentile(xs, (1 - limit) * 100)
    return [r for r in rs if r[col] >= lower and r[col] <= higher]


def isnum(*xs):
    """ Tests if x is numeric
    """
    try:
        for x in xs:
            float(x)
        return True
    except (ValueError, TypeError):
        return False


def stars(pval):
    if pval <= 0.01:
        return "***"
    elif pval <= 0.05:
        return "**"
    elif pval <= 0.10:
        return "*"
    return ""


def listify(x):
    try:
        return [x1.strip() for x1 in x.split(',')]
    except AttributeError:
        try:
            return list(iter(x))
        except TypeError:
            return [x]


def readxl(fname, sheets=None, encoding='utf-8', delimiter=None, quotechar='"', newline='\n'):
    """ Reads excel like files and yields a list of values
    """
    # locale is set in fb.run()
    def conv(x):
        try:
            return locale.atoi(x)
        except ValueError:
            try:
                return locale.atof(x)
            except Exception:
                return x
    # csv, tsv, ssv ...
    if not (fname.endswith('.xls') or fname.endswith('.xlsx')):
        # default delimiter is ","
        delimiter = delimiter or ("\t" if fname.lower().endswith('.tsv') else ",")
        with open(fname, encoding=encoding, newline=newline) as fin:
            for rs in csv.reader((x.replace('\0', '') for x in fin),\
                delimiter=delimiter, quotechar=quotechar):
                yield [conv(x) for x in rs]
    else:
        workbook = load_workbook(fname, read_only=True)
        sheets = listify(sheets) if sheets else [workbook.sheetnames[0]]
        # all sheets
        if sheets == ['*']:
            sheets = workbook.sheetnames
        for sheet in sheets:
            for row in workbook[sheet].iter_rows():
                yield [c.value for c in row]


# implicit ordering
def group(rs, key):
    keyfn = _build_keyfn(key)
    rs.sort(key=keyfn)
    return [list(rs1) for _, rs1 in groupby(rs, keyfn)]


def grouper(iterable, n, fillvalue=None):
    "Collect data into fixed-length chunks or blocks"
    # grouper('ABCDEFG', 3, 'x') --> ABC DEF Gxx
    args = [iter(iterable)] * n
    return zip_longest(fillvalue=fillvalue, *args)


def avg(rs, col, wcol=None, ndigits=None):
    if wcol:
        xs = [r for r in rs if isnum(r[col], r[wcol])]
        val = np.average([x[col] for x in xs], weights=[x[wcol] for x in xs]) if xs else ''
    else:
        xs = [r for r in rs if isnum(r[col])]
        val = np.average([x[col] for x in xs]) if xs else ''
    return round(val, ndigits) if ndigits and xs else val


def _num1(rs, c, p):
    rs = [r for r in rs if isnum(r[c])]
    rs.sort(key=lambda r: r[c])
    rss = chunk(rs, p) if isinstance(p, int) else p(rs)
    for i, rs1 in enumerate(rss, 1):
        for r in rs1:
            r['pn_' + c] = i
    return rss


def numbering(d, dep=False):
    def fni(rs, cps):
        if cps:
            c, p = cps[0]
            _num1(rs, c, p)
            fni(rs, cps[1:])
        return rs

    def fnd(rs, cps):
        if cps:
            c, p = cps[0]
            for rs1 in _num1(rs, c, p):
                fnd(rs1, cps[1:])
        return rs

    cps = [(c, p) for c, p in d.items()]
    def fn(rs):
        for c, _ in cps:
            for r in rs:
                r['pn_' + c] = ''
        return fnd(rs, cps) if dep else fni(rs, cps)
    return fn


def _empty(r):
    r0 = {}
    for c in r:
        r0[c] = ''
    return r0


def _build_keyfn(key):
    " if key is a string return a key function "
    # if the key is already a function, just return it
    if hasattr(key, '__call__'):
        return key
    colnames = listify(key)
    # special case
    if colnames == ['*']:
        return lambda r: 1

    if len(colnames) == 1:
        col = colnames[0]
        return lambda r: r[col]
    else:
        return lambda r: [r[colname] for colname in colnames]